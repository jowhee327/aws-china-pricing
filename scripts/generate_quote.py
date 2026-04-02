#!/usr/bin/env python3
"""AWS 中国区 Excel 报价单生成工具

从 CSV/Excel 工作负载生成正式的 Excel 报价单，含客户名、有效期、明细、汇总。

用法:
  python3 generate_quote.py --input workload.csv --region cn-north-1 \
    --customer "ABC公司" --output quote.xlsx

  python3 generate_quote.py --input workload.csv --region cn-north-1 \
    --customer "ABC公司" --validity 30 --include-tax \
    --discount-config ../discount-config.yaml --output quote.xlsx

  python3 generate_quote.py --input workload.csv --region cn-north-1 \
    --customer "ABC公司" --compare on-demand,ri-standard-1yr-partial --output quote.xlsx
"""

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))
from calculate_cost import (
    load_workload, load_discount_config, get_price_for_item,
    calculate_item_cost, apply_discounts,
)

# 样式定义
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=11, color="666666")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CNY_FORMAT = '#,##0.00 "CNY"'


def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row: int, values: list):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_data_row(ws, row: int, values: list, formats: list = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if formats and col <= len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]


def write_total_row(ws, row: int, values: list, formats: list = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if formats and col <= len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]


REGION_NAMES = {
    "cn-north-1": "北京 (cn-north-1)",
    "cn-northwest-1": "宁夏 (cn-northwest-1)",
    "cn-north-1-pkx-1": "Auto Cloud Local Zone (cn-north-1-pkx-1)",
}

BILLING_MODE_NAMES = {
    "on-demand": "按需",
    "ri-standard-1yr-no": "标准RI 1年 无预付",
    "ri-standard-1yr-partial": "标准RI 1年 部分预付",
    "ri-standard-1yr-all": "标准RI 1年 全预付",
    "ri-standard-3yr-no": "标准RI 3年 无预付",
    "ri-standard-3yr-partial": "标准RI 3年 部分预付",
    "ri-standard-3yr-all": "标准RI 3年 全预付",
    "ri-convertible-1yr-no": "可转换RI 1年 无预付",
    "ri-convertible-1yr-partial": "可转换RI 1年 部分预付",
    "ri-convertible-1yr-all": "可转换RI 1年 全预付",
    "ri-convertible-3yr-no": "可转换RI 3年 无预付",
    "ri-convertible-3yr-partial": "可转换RI 3年 部分预付",
    "ri-convertible-3yr-all": "可转换RI 3年 全预付",
}


def generate_quote(items: list[dict], results: list[dict], config: dict):
    """生成报价单"""
    wb = Workbook()

    # --- Sheet 1: 报价单 ---
    ws = wb.active
    ws.title = "报价单"

    customer = config.get("customer", "")
    validity = config.get("validity", 30)
    include_tax = config.get("include_tax", False)
    quote_date = datetime.now()
    expire_date = quote_date + timedelta(days=validity)
    quote_no = f"AWS-CN-{quote_date.strftime('%Y%m%d')}-{abs(hash(customer)) % 10000:04d}"

    set_col_widths(ws, [5, 15, 18, 16, 8, 10, 25, 14, 16, 16, 16, 14, 20])

    # 标题区
    ws.merge_cells("A1:M1")
    cell = ws.cell(row=1, column=1, value="AWS 中国区云服务报价单")
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 40

    info_rows = [
        ("报价编号:", quote_no, "客户名称:", customer),
        ("报价日期:", quote_date.strftime("%Y-%m-%d"), "有效期至:", expire_date.strftime("%Y-%m-%d")),
        ("税费说明:", f"{'含税 (6% 增值税)' if include_tax else '不含税'}", "货币:", "CNY (人民币)"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=l1).font = BOLD_FONT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        ws.cell(row=i, column=2, value=v1).font = NORMAL_FONT
        ws.cell(row=i, column=5, value=l2).font = BOLD_FONT
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)
        ws.cell(row=i, column=6, value=v2).font = NORMAL_FONT

    # 明细表头
    header_row = 7
    headers = [
        "序号", "服务", "实例类型", "区域", "数量", "月使用时长(h)",
        "计费模式", "小时单价(CNY)", "月费/台(CNY)", "月费合计(CNY)",
        "年费合计(CNY)", "预付金额(CNY)", "备注",
    ]
    write_header_row(ws, header_row, headers)
    ws.row_dimensions[header_row].height = 30

    # 明细数据
    data_start = header_row + 1
    total_monthly = 0
    total_yearly = 0
    total_upfront = 0
    fmt = [None, None, None, None, "#,##0", "#,##0", None,
           CNY_FORMAT, CNY_FORMAT, CNY_FORMAT, CNY_FORMAT, CNY_FORMAT, None]

    for idx, r in enumerate(results, 1):
        row_num = data_start + idx - 1
        billing_name = BILLING_MODE_NAMES.get(r.get("billing_mode", ""), r.get("billing_mode", ""))
        region_name = REGION_NAMES.get(r.get("region", ""), r.get("region", ""))
        discounts_note = ", ".join(r.get("applied_discounts", []))
        notes = r.get("notes", "")
        if discounts_note:
            notes = f"{notes} [{discounts_note}]" if notes else f"[{discounts_note}]"
        if r.get("warning"):
            notes = f"{notes} ⚠{r['warning']}" if notes else f"⚠{r['warning']}"

        values = [
            idx,
            r.get("service", ""),
            r.get("instance_type", ""),
            region_name,
            r.get("quantity", 1),
            r.get("usage_hours", 730),
            billing_name,
            r.get("hourly_after_discount", 0),
            r.get("monthly_per_unit", 0),
            r.get("monthly_total", 0),
            r.get("yearly_total", 0),
            r.get("upfront_total", 0),
            notes,
        ]
        write_data_row(ws, row_num, values, fmt)

        total_monthly += r.get("monthly_total", 0)
        total_yearly += r.get("yearly_total", 0)
        total_upfront += r.get("upfront_total", 0)

    # 合计行
    total_row = data_start + len(results)
    total_values = [
        "", "合计", "", "", "", "", "",
        "", "", total_monthly, total_yearly, total_upfront, "",
    ]
    write_total_row(ws, total_row, total_values, fmt)

    # 注脚
    foot_row = total_row + 2
    notes_text = [
        "说明:",
        "1. 以上报价基于 AWS 中国区公开定价（list price），实际费用以 AWS 账单为准。",
        "2. 预留实例(RI)和 Savings Plans 需要签订承诺合同。",
        "3. 数据传输费、请求费等额外费用未包含在实例费用中，请另行计算。",
        f"4. 本报价有效期至 {expire_date.strftime('%Y年%m月%d日')}。",
    ]
    if include_tax:
        notes_text.append("5. 以上金额已包含 6% 增值税。")
    for i, note in enumerate(notes_text):
        cell = ws.cell(row=foot_row + i, column=1, value=note)
        cell.font = Font(name="微软雅黑", size=9, color="666666")
        ws.merge_cells(start_row=foot_row + i, start_column=1, end_row=foot_row + i, end_column=13)

    # --- Sheet 2: 费率对比（如果有 RI 数据） ---
    has_ri = any(r.get("billing_mode", "").startswith("ri-") for r in results)
    if has_ri or config.get("compare"):
        ws2 = wb.create_sheet("计费模式对比")
        ws2.cell(row=1, column=1, value="各计费模式成本对比").font = TITLE_FONT
        ws2.merge_cells("A1:H1")
        # Placeholder for comparison data — populated by compare mode
        ws2.cell(row=3, column=1, value="详细对比数据请使用 --compare 参数生成").font = SUBTITLE_FONT

    return wb


def main():
    parser = argparse.ArgumentParser(description="AWS 中国区 Excel 报价单生成工具")
    parser.add_argument("--input", "-i", required=True, help="输入文件 (CSV 或 Excel)")
    parser.add_argument("--region", "-r", default="cn-north-1", help="默认区域")
    parser.add_argument("--customer", default="", help="客户名称")
    parser.add_argument("--validity", type=int, default=30, help="报价有效期（天）")
    parser.add_argument("--output", "-o", default="quote.xlsx", help="输出 Excel 文件路径")
    parser.add_argument("--discount-config", "-d",
                       default=str(SCRIPT_DIR.parent / "discount-config.yaml"),
                       help="折扣配置文件路径")
    parser.add_argument("--include-tax", action="store_true", help="含 6% 增值税")
    parser.add_argument("--compare", "-c", help="对比计费模式（逗号分隔）")
    args = parser.parse_args()

    # 加载配置
    discount_config = load_discount_config(args.discount_config)

    # 加载工作负载
    items = load_workload(args.input)
    if not items:
        print("错误: 输入文件为空", file=sys.stderr)
        sys.exit(1)

    print(f"已加载 {len(items)} 个条目，正在查询价格...", file=sys.stderr)

    # 设置默认区域
    for item in items:
        if not item.get("region"):
            item["region"] = args.region

    # 计算成本
    results = []
    for item in items:
        billing_mode = item.get("billing_mode", "on-demand") or "on-demand"
        print(f"  查询 {item.get('service', '')} {item.get('instance_type', '')} ...", file=sys.stderr)
        price_data = get_price_for_item(item, billing_mode=billing_mode)
        if not price_data:
            results.append({
                "service": item.get("service", ""),
                "instance_type": item.get("instance_type", ""),
                "region": item.get("region", ""),
                "quantity": int(item.get("quantity", 1) or 1),
                "usage_hours": float(item.get("usage_hours", 730) or 730),
                "billing_mode": billing_mode,
                "hourly_list": 0, "hourly_after_discount": 0,
                "monthly_per_unit": 0, "monthly_total": 0,
                "upfront_total": 0, "yearly_total": 0,
                "warning": "未找到价格",
                "applied_discounts": [],
                "notes": item.get("notes", ""),
                "currency": "CNY",
            })
            continue
        cost = calculate_item_cost(item, price_data, discount_config, args.include_tax)
        results.append(cost)

    # 生成报价单
    gen_config = {
        "customer": args.customer,
        "validity": args.validity,
        "include_tax": args.include_tax,
        "compare": args.compare,
    }
    wb = generate_quote(items, results, gen_config)

    # 保存
    wb.save(args.output)
    print(f"\n报价单已生成: {args.output}", file=sys.stderr)

    # 打印摘要
    total_monthly = sum(r.get("monthly_total", 0) for r in results)
    total_yearly = sum(r.get("yearly_total", 0) for r in results)
    tax_label = "（含税）" if args.include_tax else "（不含税）"
    print(f"\n摘要{tax_label}:", file=sys.stderr)
    print(f"  月费用合计: CNY {total_monthly:,.2f}", file=sys.stderr)
    print(f"  年费用合计: CNY {total_yearly:,.2f}", file=sys.stderr)
    if args.customer:
        print(f"  客户: {args.customer}", file=sys.stderr)


if __name__ == "__main__":
    main()
