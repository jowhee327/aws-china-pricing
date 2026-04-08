#!/usr/bin/env python3
"""AWS 中国区智能工作负载导入工具 v2

支持任意格式 Excel 文件：
- 多 Sheet 支持（每个 sheet 独立处理）
- 智能列检测（语义关键词匹配）
- 智能行解析（标题行/表头行/数据行自动识别）
- 增强服务映射（更多中文别名 + 应用层服务 fallback）
- 规格智能提取（多格式支持）
- 数量智能处理（公式、按量付费等）
- 默认直接输出 Excel 报价单（smart_import → calculate_cost → generate_quote 一步到位）

用法:
  # 最简用法：输入 Excel，输出 Excel 报价单（默认输出 {输入文件名}_报价单.xlsx）
  python3 smart_import.py --input LBS.xlsx --region cn-north-1

  # 指定输出 Excel 报价单
  python3 smart_import.py --input LBS.xlsx --output quote.xlsx --region cn-north-1

  # 输出标准化 CSV（保留旧行为）
  python3 smart_import.py --input raw_workload.xlsx --output standardized.csv --region cn-north-1

  # 终端输出模式（调用 calculate_cost.py）
  python3 smart_import.py --input raw_workload.csv --region cn-north-1 --calculate
"""

import argparse
import csv
import json
import re
import subprocess
import sys
from pathlib import Path

# ── 服务映射表（增强版）────────────────────────────────────────────────────
# 每条规则: (关键词列表, ServiceCode, 附加字段 dict)
# 关键词全部小写，匹配时也先 lower()
SERVICE_RULES: list[tuple[list[str], str, dict]] = [
    # ── 计算类 ──
    (["compute", "计算", "服务器", "虚拟机", "vm", "server", "ec2", "ecs",
      "云服务器"], "AmazonEC2", {}),
    (["container", "容器", "fargate", "ecs容器", "docker"], "AmazonECS", {}),
    (["k8s", "kubernetes", "eks"], "AmazonEKS", {}),
    (["ecr", "容器注册", "container registry", "镜像仓库"], "AmazonECR", {}),
    (["serverless", "无服务器", "lambda", "函数计算", "function"], "AWSLambda", {}),

    # ── 数据库类（具体引擎优先）──
    (["documentdb", "文档数据库", "mongodb"], "AmazonDocDB", {}),
    (["neptune", "图数据库", "graph"], "AmazonNeptune", {}),
    (["timestream", "时序数据库", "timeseries"], "AmazonTimestream", {}),
    (["dynamodb", "nosql", "键值数据库", "key-value"], "AmazonDynamoDB", {}),
    (["redshift", "数仓", "数据仓库", "data warehouse"], "AmazonRedshift", {}),
    (["memorydb", "内存数据库"], "AmazonMemoryDB", {}),
    (["aurora mysql", "aurora-mysql"], "AmazonRDS", {"engine": "Aurora MySQL"}),
    (["aurora postgresql", "aurora-postgresql", "aurora postgres"],
     "AmazonRDS", {"engine": "Aurora PostgreSQL"}),
    (["aurora"], "AmazonRDS", {"engine": "Aurora MySQL"}),
    (["mysql", "mysql数据库"], "AmazonRDS", {"engine": "MySQL"}),
    (["postgresql", "postgres", "pg数据库"], "AmazonRDS", {"engine": "PostgreSQL"}),
    (["mariadb"], "AmazonRDS", {"engine": "MariaDB"}),
    (["oracle", "oracle数据库"], "AmazonRDS", {"engine": "Oracle"}),
    (["sql server", "sqlserver", "mssql"], "AmazonRDS", {"engine": "SQL Server"}),
    (["rds", "database", "数据库", "关系数据库"], "AmazonRDS", {"engine": "MySQL"}),

    # ── 缓存类 ──
    (["dax", "dynamodb加速"], "AmazonDAX", {}),
    (["valkey"], "AmazonElastiCache", {"engine": "Valkey"}),
    (["memcached"], "AmazonElastiCache", {"engine": "Memcached"}),
    (["redis"], "AmazonElastiCache", {"engine": "Redis"}),
    (["缓存", "elasticache", "cache"], "AmazonElastiCache", {"engine": "Redis"}),

    # ── 存储类 ──
    (["glacier", "归档", "冷存储", "archive"], "AmazonGlacier", {}),
    (["fsx", "windows文件", "lustre"], "AmazonFSx", {}),
    (["efs", "文件存储", "弹性文件服务", "弹性文件", "nfs", "nas",
      "file storage"], "AmazonEFS", {}),
    (["ebs", "块存储", "弹性块存储", "弹性块", "ssd存储", "通用ssd", "通用SSD",
      "云盘", "block storage"], "AmazonEBS", {"productFamily": "Storage",
                                              "volumeApiName": "gp3"}),
    (["对象存储", "s3", "oss", "object storage"], "AmazonS3", {}),

    # ── 网络类 ──
    (["cdn", "加速", "cloudfront", "内容分发"], "AmazonCloudFront", {}),
    (["nlb", "network load balancer", "网络负载均衡"], "AWSELB", {"_lb_type": "NLB"}),
    (["clb", "classic load balancer", "经典负载均衡"], "AWSELB", {"_lb_type": "CLB"}),
    (["alb", "应用负载均衡", "application load balancer"], "AWSELB", {"_lb_type": "ALB"}),
    (["负载均衡", "lb", "slb", "elb", "load balancer"], "AWSELB", {"_lb_type": "ALB"}),
    (["专线", "direct connect", "直连"], "AWSDirectConnect", {}),
    (["dns", "route53", "域名"], "AMAZONROUTE53REGIONALCHINA", {}),
    (["waf", "web防火墙"], "awswaf", {}),
    (["network firewall", "网络防火墙"], "AWSNetworkFirewall", {}),
    (["vpc", "vpn", "虚拟专用网络", "vpcendpoint",
      "vpc endpoint"], "AmazonVPC", {}),

    # ── 消息/流式 ──
    (["sqs", "消息队列", "队列", "message queue"], "AWSQueueService", {}),
    (["sns", "通知", "推送", "notification"], "AmazonSNS", {}),
    (["kafka", "msk", "消息流"], "AmazonMSK", {}),
    (["rabbitmq"], "AmazonMQ", {"engine": "RabbitMQ"}),
    (["activemq"], "AmazonMQ", {"engine": "ActiveMQ"}),
    (["mq", "消息代理"], "AmazonMQ", {"engine": "RabbitMQ"}),
    (["eventbridge", "事件"], "AWSEvents", {}),
    (["kinesis firehose", "数据投递"], "AmazonKinesisFirehose", {}),
    (["kinesis analytics", "流分析"], "AmazonKinesisAnalytics", {}),
    (["kinesis video", "视频流"], "AmazonKinesisVideo", {}),
    (["kinesis", "实时流", "streaming"], "AmazonKinesis", {}),

    # ── 大数据/分析 ──
    (["emr", "spark", "mapreduce", "大数据"], "ElasticMapReduce", {}),
    (["athena", "sql查询", "交互式查询"], "AmazonAthena", {}),
    (["glue", "etl", "数据集成"], "AWSGlue", {}),
    (["quicksight", "bi", "报表", "可视化"], "AmazonQuickSight", {}),

    # ── AI/ML ──
    (["sagemaker", "机器学习", "ml", "ai"], "AmazonSageMaker", {}),
    (["polly", "语音合成", "tts"], "AmazonPolly", {}),
    (["transcribe", "语音转文字", "str", "asr"], "transcribe", {}),
    (["personalize", "推荐"], "AmazonPersonalize", {}),

    # ── 安全类 ──
    (["kms", "密钥管理", "加密"], "awskms", {}),
    (["acm", "证书", "ssl", "https"], "ACM", {}),
    (["guardduty", "安全审计", "威胁检测"], "AmazonGuardDuty", {}),
    (["inspector", "漏洞扫描"], "AmazonInspectorV2", {}),
    (["secrets manager", "密钥存储", "密码管理"], "AWSSecretsManager", {}),
    (["security hub", "安全中心"], "AWSSecurityHub", {}),
    (["firewall manager", "防火墙管理"], "AWSFMS", {}),
    (["iam access analyzer"], "AWSIAMAccessAnalyzer", {}),

    # ── 运维/管理 ──
    (["cloudwatch", "监控", "告警", "日志服务", "云监控"], "AmazonCloudWatch", {}),
    (["cloudtrail", "日志审计"], "AWSCloudTrail", {}),
    (["backup", "备份"], "AWSBackup", {}),
    (["workspaces", "桌面", "云桌面", "vdi"], "AmazonWorkSpaces", {}),
    (["dms", "数据迁移", "数据库迁移"], "AWSDatabaseMigrationSvc", {}),
    (["datasync", "数据同步"], "AWSDataSync", {}),
    (["snowball", "数据搬迁"], "IngestionServiceSnowball", {}),
    (["iot analytics"], "AWSIoTAnalytics", {}),
    (["iot events"], "AWSIoTEvents", {}),
    (["iot sitewise"], "AWSIoTSiteWise", {}),
    (["iot", "物联网"], "AWSIoT", {}),
    (["step functions", "工作流", "状态机"], "AmazonStates", {}),
    (["systems manager", "运维管理"], "AWSSystemsManager", {}),
    (["config", "配置审计"], "AWSConfig", {}),
    (["service catalog", "服务目录"], "AWSServiceCatalog", {}),
    (["cloudformation", "iac", "基础设施即代码"], "AWSCloudFormation", {}),
    (["codebuild", "构建"], "CodeBuild", {}),
    (["codecommit", "代码仓库"], "AWSCodeCommit", {}),
    (["codedeploy", "部署"], "AWSCodeDeploy", {}),
    (["codepipeline", "流水线", "ci/cd"], "AWSCodePipeline", {}),
    (["x-ray", "链路追踪", "tracing"], "AWSXRay", {}),
    (["appsync", "graphql"], "AWSAppSync", {}),
    (["api gateway", "api网关"], "AmazonApiGateway", {}),
    (["transfer family", "sftp", "ftp"], "AWSTransfer", {}),
    (["greengrass", "边缘计算"], "AWSGreengrass", {}),
    (["storage gateway", "混合存储"], "AWSStorageGateway", {}),
    (["mediaconvert", "媒体转码"], "AWSElementalMediaConvert", {}),
    (["gamelift", "游戏服务器"], "AmazonGameLift", {}),
    (["cost explorer", "成本分析"], "AWSCostExplorer", {}),
    (["budgets", "预算"], "AWSBudgets", {}),
    (["compute optimizer", "优化建议"], "AWSComputeOptimizer", {}),
    (["cloudmap", "服务发现"], "AWSCloudMap", {}),
    (["verified permissions", "授权"], "AmazonVerifiedPermissions", {}),
    (["mwaa", "airflow", "工作流调度"], "AmazonMWAA", {}),
    (["opensearch", "elasticsearch", "搜索", "es"], "AmazonES", {}),
    (["swf", "简单工作流"], "AmazonSWF", {}),

    # ── 应用层服务（通常运行在 EC2 实例上）──
    (["脱密服务", "脱密", "脱敏服务", "脱敏"], "AmazonEC2", {}),
    (["eureka"], "AmazonEC2", {}),
    (["nacos"], "AmazonEC2", {}),
    (["网关服务"], "AmazonEC2", {}),
    (["portal服务", "protal服务", "portal", "protal", "门户"], "AmazonEC2", {}),
    (["自研监控", "monitor服务"], "AmazonEC2", {}),
    (["收容服务"], "AmazonEC2", {}),
]

# 引擎提示：从输入文本中推断 engine / cacheEngine
ENGINE_HINTS = {
    "aurora mysql": ("engine", "Aurora MySQL"),
    "aurora postgresql": ("engine", "Aurora PostgreSQL"),
    "aurora postgres": ("engine", "Aurora PostgreSQL"),
    "aurora": ("engine", "Aurora MySQL"),
    "mysql": ("engine", "MySQL"),
    "postgresql": ("engine", "PostgreSQL"),
    "postgres": ("engine", "PostgreSQL"),
    "mariadb": ("engine", "MariaDB"),
    "oracle": ("engine", "Oracle"),
    "sql server": ("engine", "SQL Server"),
    "redis": ("cacheEngine", "Redis"),
    "memcached": ("cacheEngine", "Memcached"),
    "valkey": ("cacheEngine", "Valkey"),
}

# EBS 卷类型提示词
EBS_VOLUME_HINTS: list[tuple[list[str], str]] = [
    (["gp3"], "gp3"),
    (["gp2"], "gp2"),
    (["io2", "预置iops", "provisioned iops"], "io2"),
    (["io1"], "io1"),
    (["st1", "吞吐优化"], "st1"),
    (["sc1", "冷hdd", "cold hdd"], "sc1"),
    (["通用ssd", "通用SSD", "通用型ssd", "ssd存储", "快速存储"], "gp3"),
]


def detect_ebs_volume_type(text: str) -> str:
    """从文本中检测 EBS 卷类型，默认 gp3"""
    text_lower = text.lower()
    for keywords, volume_type in EBS_VOLUME_HINTS:
        for kw in keywords:
            if kw in text_lower:
                return volume_type
    return "gp3"


# 非AWS标准服务替代建议
NON_STANDARD_SERVICE_SUGGESTIONS: dict[str, str] = {
    "云堡垒机": "建议使用 AWS Systems Manager Session Manager",
    "堡垒机": "建议使用 AWS Systems Manager Session Manager",
}

# 需要实例规格的服务（缺少规格时应提醒用户）
SERVICES_NEED_SPEC = {
    "AmazonEC2", "AmazonRDS", "AmazonElastiCache", "AmazonDocDB",
    "AmazonNeptune", "AmazonRedshift", "AmazonMemoryDB", "AmazonMQ",
    "AmazonES", "ElasticMapReduce",
}

# ── 托管服务 Graviton 实例映射 ────────────────────────────────────────────
# EC2 保持 Intel，以下托管服务默认推荐最新代 Graviton (r7g/m7g)
MANAGED_GRAVITON_SERVICES = {
    "AmazonRDS", "AmazonDocDB", "AmazonNeptune",
    "AmazonMemoryDB", "AmazonElastiCache", "AmazonES",
}

# 托管服务实例前缀
MANAGED_SERVICE_PREFIX = {
    "AmazonRDS": "db",
    "AmazonDocDB": "db",
    "AmazonNeptune": "db",
    "AmazonMemoryDB": "db",
    "AmazonElastiCache": "cache",
    "AmazonES": "",  # OpenSearch: {family}.{size}.search
}

# Graviton r7g (memory-optimized) 规格表: (vcpu, memory_gib, size_name)
GRAVITON_R7G_SPECS = [
    (2, 16, "large"),
    (4, 32, "xlarge"),
    (8, 64, "2xlarge"),
    (16, 128, "4xlarge"),
    (32, 256, "8xlarge"),
    (48, 384, "12xlarge"),
    (64, 512, "16xlarge"),
]

# Graviton m7g (general-purpose) 规格表: (vcpu, memory_gib, size_name)
GRAVITON_M7G_SPECS = [
    (2, 8, "large"),
    (4, 16, "xlarge"),
    (8, 32, "2xlarge"),
    (16, 64, "4xlarge"),
    (32, 128, "8xlarge"),
    (48, 192, "12xlarge"),
    (64, 256, "16xlarge"),
]

# Graviton r6g (memory-optimized, 上一代) — DocumentDB 中国区仅支持 r6g
GRAVITON_R6G_SPECS = [
    (2, 16, "large"),
    (4, 32, "xlarge"),
    (8, 64, "2xlarge"),
    (16, 128, "4xlarge"),
    (32, 256, "8xlarge"),
    (48, 384, "12xlarge"),
    (64, 512, "16xlarge"),
]

def normalize_billing_mode(mode: str) -> str:
    """将简化的计费模式名称标准化为完整名称"""
    # 处理简化名称：ri-1y-no-upfront -> ri-standard-1yr-no
    if mode.startswith("ri-") and not mode.startswith("ri-sp-") and not mode.startswith("ri-standard-") and not mode.startswith("ri-convertible-"):
        # 默认为 standard RI
        mode = mode.replace("ri-", "ri-standard-")
    if mode.startswith("sp-") and not mode.startswith("sp-compute-") and not mode.startswith("sp-instance-"):
        # 默认为 compute SP
        mode = mode.replace("sp-", "sp-compute-")
    # 标准化年份表示：1y -> 1yr, 3y -> 3yr
    return mode.replace("1y-", "1yr-").replace("3y-", "3yr-")


# 中国区各托管服务实际可用的 Graviton 实例族
# DocumentDB: 仅 r6g（无 r7g/m7g/m6g）
# ElastiCache: r7g + m7g (xlarge+，无 m7g.large)
# 其他: r7g + m7g
MANAGED_GRAVITON_FAMILY: dict[str, dict[str, tuple[str, list]]] = {
    "AmazonDocDB": {
        "r": ("r6g", GRAVITON_R6G_SPECS),
        "m": ("r6g", GRAVITON_R6G_SPECS),  # DocDB 无 m 系列 Graviton，用 r6g
    },
}


# ── 列角色检测关键词 ──────────────────────────────────────────────────────
COLUMN_ROLE_KEYWORDS = {
    "service_type": ["类型", "云产品", "产品分类", "服务类型", "资源类型",
                     "type", "service", "云产品分类"],
    "spec": ["配置", "规格", "详情", "spec", "configuration", "config", "名称"],
    "quantity": ["数量", "个数", "台数", "count", "quantity", "num",
                 "用量", "最低个数"],
    "unit": ["单位", "unit"],
    "description": ["描述", "备注", "说明", "用途", "description", "note",
                     "使用场景"],
    "business": ["业务场景", "业务分类", "业务", "business"],
}

# 扁平化关键词集合（用于快速表头行检测）
_ALL_HEADER_KEYWORDS: set[str] = set()
for _kws in COLUMN_ROLE_KEYWORDS.values():
    _ALL_HEADER_KEYWORDS.update(_kws)

# CSV 输入的列名别名（兼容旧格式）
COLUMN_ALIASES = {
    "类型": "service", "服务": "service", "service": "service", "服务类型": "service",
    "规格": "spec", "配置": "spec", "spec": "spec", "实例": "spec", "instance": "spec",
    "instance_type": "instance_type",
    "数量": "quantity", "qty": "quantity", "quantity": "quantity", "台数": "quantity",
    "备注": "notes", "note": "notes", "notes": "notes", "说明": "notes", "描述": "notes",
    "region": "region", "区域": "region",
    "os": "os", "操作系统": "os",
    "engine": "engine", "引擎": "engine",
    "storage_gb": "storage_gb", "存储": "storage_gb",
    "billing_mode": "billing_mode", "计费模式": "billing_mode",
    "usage_hours": "usage_hours", "使用时长": "usage_hours",
}

# 标准输出列
OUTPUT_FIELDS = [
    "sheet_name", "service", "instance_type", "region", "quantity", "usage_hours",
    "os", "engine", "storage_gb", "billing_mode", "productFamily", "volumeApiName",
    "notes", "original_request", "section",
]


# ── 核心匹配逻辑 ──────────────────────────────────────────────────────────

def match_service(text: str) -> list[tuple[str, str, dict, int]]:
    """对输入文本进行服务匹配，返回 [(matched_keyword, service_code, extra_fields, score)] 按 score 降序"""
    text_lower = text.lower()
    matches = []

    for keywords, service_code, extra in SERVICE_RULES:
        best_kw = None
        best_score = 0
        for kw in keywords:
            kw_lower = kw.lower()
            if kw_lower in text_lower:
                score = len(kw_lower) * 10
                if text_lower.strip() == kw_lower:
                    score += 100
                if score > best_score:
                    best_score = score
                    best_kw = kw
        if best_kw:
            matches.append((best_kw, service_code, extra, best_score))

    # 去重同一 service_code，保留最高分
    seen = {}
    for kw, sc, extra, score in matches:
        if sc not in seen or score > seen[sc][3]:
            seen[sc] = (kw, sc, extra, score)

    return sorted(seen.values(), key=lambda x: -x[3])


def match_service_smart(text: str) -> tuple[str, dict, str, str]:
    """智能服务匹配（支持复合条目拆分）。
    返回 (service_code, extra_fields, warning, matched_keyword)"""
    if not text or not text.strip():
        return "", {}, "无服务描述", ""

    # 如果文本有逗号/顿号等分隔符，先尝试按第一个部分匹配
    parts = re.split(r'[,，、]+', text)
    if len(parts) > 1:
        first = parts[0].strip()
        if first:
            m = match_service(first)
            if m:
                _, sc, extra, _ = m[0]
                others = ", ".join(p.strip() for p in parts[1:] if p.strip())
                warn = f"复合条目，主服务={sc}，其余: {others}" if others else ""
                return sc, extra, warn, m[0][0]

    # 整体匹配
    matches = match_service(text)
    if not matches:
        return "", {}, "", ""
    if len(matches) == 1:
        _, sc, extra, _ = matches[0]
        return sc, extra, "", matches[0][0]
    # 多匹配：取最高分
    top, second = matches[0], matches[1]
    _, sc, extra, _ = top
    if top[3] > second[3]:
        return sc, extra, "", top[0]
    alternatives = ", ".join(f"{m[1]}('{m[0]}')" for m in matches[:3])
    return sc, extra, f"多个匹配: {alternatives}，已选择 {sc}", top[0]


def extract_spec(text: str) -> dict:
    """从规格描述中提取 vCPU、内存、存储等信息（增强版）"""
    info: dict = {}
    if not text:
        return info

    # 1. CPU+Memory: "8C16G", "4C/16G", "16C/32G", "8核16G", "12c32g500G"
    m = re.search(r'(\d+)\s*[cC核vV]\s*[pP]?[uU]?\s*/?\s*(\d+)\s*[gG]', text)
    if m:
        info["vcpu"] = int(m.group(1))
        info["memory"] = int(m.group(2))
        # 检查紧跟的存储: "12c32g500G"
        rest = text[m.end():]
        sm = re.match(r'(\d+)\s*[gG]', rest)
        if sm:
            info["storage_gb"] = int(sm.group(1))

    # 2. 独立内存 (没有 CPU 信息时): "8G", "8G-Cluster", "4G内存"
    if "memory" not in info:
        for mm in re.finditer(
            r'(\d+)\s*[gG][iI]?[bB]?\s*(?:[-\s]|内存|memory|cluster|$)',
            text, re.IGNORECASE
        ):
            start = mm.start()
            prefix = text[max(0, start - 3):start]
            if not re.search(r'存储|storage|磁盘|disk', prefix, re.IGNORECASE):
                info["memory"] = int(mm.group(1))
                break

    # 3. 存储: "存储128G", "1TB", "500G存储"
    if "storage_gb" not in info:
        m = re.search(r'(\d+)\s*[tT][bB]', text)
        if m:
            info["storage_gb"] = int(m.group(1)) * 1024
        else:
            m = re.search(r'(?:存储|storage|磁盘|disk)\s*(\d+)\s*[gG]',
                          text, re.IGNORECASE)
            if not m:
                m = re.search(r'(\d+)\s*[gG][bB]?\s*(?:存储|storage|磁盘|disk)',
                              text, re.IGNORECASE)
            if m:
                info["storage_gb"] = int(m.group(1))

    # 4. 带宽/IO 标注
    notes = []
    m = re.search(r'\d+\s*[mM][bB]/s\S*', text)
    if m:
        notes.append(m.group(0))
    m = re.search(r'\d+\s*[mMgG]\s*带宽', text)
    if m:
        notes.append(m.group(0))
    if notes:
        info["spec_notes"] = "; ".join(notes)

    return info


def detect_engine(text: str) -> dict:
    """从文本中检测数据库/缓存引擎"""
    text_lower = text.lower()
    for hint, (field, value) in ENGINE_HINTS.items():
        if hint in text_lower:
            return {field: value}
    return {}


def parse_quantity(value) -> tuple[int, str]:
    """解析数量，返回 (数量, 备注)。

    支持: 纯数字、Excel 公式、按量付费、非数字字符串
    """
    if value is None:
        return 1, ""

    s = str(value).strip()
    if not s or s.lower() == "none":
        return 1, ""

    # 纯数字
    try:
        n = float(s)
        return max(1, int(n)), ""
    except (ValueError, OverflowError):
        pass

    # Excel 公式
    if s.startswith("="):
        return 1, f"公式: {s}"

    # "按量付费"
    if "按量" in s:
        return 1, "按量付费"

    # 带单位的整数: "3台", "2个"
    m = re.match(r'^(\d+)\s*[台个实例件套]', s)
    if m:
        return int(m.group(1)), ""

    # 其他文本 → 默认 1，原文作为备注（可能含规格信息）
    return 1, s


# ── 智能行分类 ──────────────────────────────────────────────────────────────

def _is_header_keyword(text: str) -> bool:
    """检查文本是否包含任意列检测关键词"""
    text_lower = text.lower()
    return any(kw in text_lower for kw in _ALL_HEADER_KEYWORDS)


def classify_row(row_values: list, has_col_map: bool = False) -> tuple[str, dict]:
    """将一行分类为 empty / title / header / data。

    Args:
        row_values: 行数据列表
        has_col_map: 是否已有列映射。为 True 时跳过表头检测，优先当 data 处理。

    返回: (分类, 信息字典)
    """
    non_empty = [(i, v) for i, v in enumerate(row_values)
                 if v is not None and str(v).strip()]

    if not non_empty:
        return "empty", {}

    n = len(non_empty)

    # 标题行: 仅 1 个非空单元格，长文本，非纯数字
    if n == 1:
        val = str(non_empty[0][1]).strip()
        if len(val) > 3 and not re.match(r'^[\d.,]+$', val):
            return "title", {"section": val}
        return "data", {}

    # 已有列映射时跳过表头检测，优先当 data 处理
    if has_col_map:
        return "data", {}

    # 表头行: ≥2 个非空单元格, ≥2 个匹配关键词, 无纯数字单元格
    if n >= 2:
        n_keyword = 0
        n_numeric = 0
        for _, v in non_empty:
            s = str(v).strip()
            if _is_header_keyword(s):
                n_keyword += 1
            if re.match(r'^[\d.,]+$', s):
                n_numeric += 1
        if n_keyword >= 2 and n_numeric == 0:
            return "header", {}

    return "data", {}


def detect_column_roles(header_values: list) -> dict[int, str]:
    """根据表头值检测每列的角色。返回 {列索引: 角色名}"""
    roles = {}
    for i, v in enumerate(header_values):
        if v is None or not str(v).strip():
            continue
        text = str(v).strip().lower()
        best_role = None
        best_score = 0
        for role, keywords in COLUMN_ROLE_KEYWORDS.items():
            for kw in keywords:
                if kw in text:
                    score = len(kw)
                    if score > best_score:
                        best_score = score
                        best_role = role
        if best_role:
            roles[i] = best_role
        else:
            roles[i] = f"col_{i}"
    return roles


def merge_header_rows(row1: list, row2: list) -> list:
    """合并两行表头（处理跨行表头/子表头）"""
    n = max(len(row1), len(row2))
    merged = [None] * n
    for i in range(n):
        v1 = row1[i] if i < len(row1) else None
        v2 = row2[i] if i < len(row2) else None
        s1 = str(v1).strip() if v1 is not None else ""
        s2 = str(v2).strip() if v2 is not None else ""
        if s1 and s2:
            merged[i] = f"{s1}/{s2}"
        elif s2:
            merged[i] = v2
        else:
            merged[i] = v1
    return merged


# ── Excel 多 Sheet 处理 ────────────────────────────────────────────────────

def process_sheet(ws, sheet_name: str, region: str, billing_mode: str = "on-demand") -> list[dict]:
    """处理单个 Excel Sheet，返回标准化条目列表"""
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return []

    n_cols = max(len(r) for r in rows_raw)
    rows = [list(r) + [None] * (n_cols - len(r)) for r in rows_raw]

    # 状态机处理（动态分类，依据当前列映射状态决定是否检测表头）
    items = []
    current_section = ""
    current_roles: dict[int, str] | None = None

    i = 0
    while i < len(rows):
        cls, info = classify_row(rows[i], has_col_map=current_roles is not None)

        if cls == "empty":
            i += 1
            continue

        if cls == "title":
            current_section = info["section"]
            i += 1
            continue

        if cls == "header":
            header_vals = rows[i]
            # 检查下一行是否也是表头（多行表头合并）
            if i + 1 < len(rows):
                next_cls, _ = classify_row(rows[i + 1], has_col_map=False)
                if next_cls == "header":
                    header_vals = merge_header_rows(rows[i], rows[i + 1])
                    i += 1
            current_roles = detect_column_roles(header_vals)
            i += 1
            continue

        if cls == "data":
            if current_roles:
                item = build_item(
                    rows[i], current_roles, sheet_name, current_section, region, billing_mode
                )
                if item:
                    items.append(item)
            else:
                print(f"Sheet '{sheet_name}' 第 {i+1} 行: "
                      f"数据行出现在表头之前，已跳过", file=sys.stderr)
            i += 1
            continue

        i += 1

    return items


def build_item(row_values: list, column_roles: dict[int, str],
               sheet_name: str, section: str, region: str, billing_mode: str = "on-demand") -> dict | None:
    """根据列角色映射，从数据行构建标准化条目"""
    # 按角色收集值（同角色多列合并）
    role_values: dict[str, str] = {}
    for col_idx, role in column_roles.items():
        if col_idx < len(row_values) and row_values[col_idx] is not None:
            val = str(row_values[col_idx]).strip()
            if val and val.lower() != "none":
                if role in role_values:
                    role_values[role] = role_values[role] + " " + val
                else:
                    role_values[role] = val

    # 提取各角色的原始值
    service_text = role_values.get("service_type", "")

    # 跳过重复表头行（service 值完全匹配列角色关键词）
    if service_text and service_text.lower() in _ALL_HEADER_KEYWORDS:
        return None

    spec_text = role_values.get("spec", "")
    quantity_text = role_values.get("quantity", "")
    unit_text = role_values.get("unit", "")
    description_text = role_values.get("description", "")
    business_text = role_values.get("business", "")

    # 跳过完全空行
    all_text = " ".join(filter(None, [
        service_text, spec_text, quantity_text, description_text, business_text
    ]))
    if not all_text.strip():
        return None

    # 保留原始行内容
    original_parts = [str(v).strip() for v in row_values
                      if v is not None and str(v).strip() and str(v).lower() != "none"]
    original_request = " | ".join(original_parts)

    # ── 服务匹配 ──
    service_code, extra_fields, warning, _ = match_service_smart(service_text)
    if not service_code:
        service_code, extra_fields, warning, _ = match_service_smart(all_text)

    if not service_code:
        # 最终 fallback: 有规格信息则视为 EC2 应用服务
        fallback_spec_text = " ".join(filter(None, [spec_text, quantity_text]))
        fallback_spec = extract_spec(fallback_spec_text)
        if fallback_spec.get("vcpu") or fallback_spec.get("memory"):
            service_code = "AmazonEC2"
            warning = f"应用服务 '{service_text}' 映射为EC2实例"
        else:
            service_code = service_text.strip() if service_text.strip() else "UNKNOWN"
            # 检查是否有已知替代建议
            suggestion = None
            for key, sug in NON_STANDARD_SERVICE_SUGGESTIONS.items():
                if key in service_text:
                    suggestion = sug
                    break
            if suggestion:
                warning = f"⚠️ 非AWS标准服务，无法自动查价。{suggestion}"
            else:
                warning = f"⚠️ 非AWS标准服务，请手动补充对应AWS服务"

    # ── 数量解析 ──
    qty, qty_note = parse_quantity(quantity_text)

    # ── 规格提取（合并多列信息）──
    all_spec_text = " ".join(filter(None, [spec_text, qty_note]))
    spec_info = extract_spec(all_spec_text)
    for t in [service_text, description_text]:
        for k, v in extract_spec(t).items():
            if k not in spec_info:
                spec_info[k] = v

    # ── 单位处理（GiB/GB/TB 表示存储大小，不是实例数量）──
    unit_clean = unit_text.lower().rstrip("/月")
    if unit_clean in ("gib", "gb", "g"):
        is_storage = (extra_fields.get("productFamily") == "Storage"
                      or service_code in ("AmazonEFS", "AmazonS3", "AmazonFSx"))
        if is_storage:
            spec_info.setdefault("storage_gb", qty)
            qty = 1
    elif unit_clean in ("tb", "t"):
        is_storage = (extra_fields.get("productFamily") == "Storage"
                      or service_code in ("AmazonEFS", "AmazonS3", "AmazonFSx"))
        if is_storage:
            spec_info.setdefault("storage_gb", qty * 1024)
            qty = 1

    # ── 引擎检测 ──
    engine_info = detect_engine(all_text)

    # ── 计费模式 ──
    # 混合模式支持：ri-sp 表示 EC2 用 SP，其他服务用 RI
    item_billing_mode = billing_mode
    if billing_mode.startswith("ri-sp-"):
        # 提取基础配置：ri-sp-1y-no-upfront -> 1y-no-upfront
        base_config = billing_mode.replace("ri-sp-", "")
        if service_code == "AmazonEC2":
            # EC2 使用对应的 SP
            item_billing_mode = normalize_billing_mode(f"sp-{base_config}")
        else:
            # 其他服务使用 RI
            item_billing_mode = normalize_billing_mode(f"ri-{base_config}")
    else:
        # 非混合模式：标准化简化名称
        item_billing_mode = normalize_billing_mode(billing_mode)

    # ── 构建备注（仅系统生成信息，不重复原始需求）──
    notes_parts: list[str] = []
    if warning:
        notes_parts.append(f"{warning}")
    if qty_note and qty_note != "按量付费":
        notes_parts.append(qty_note)
    if qty_note == "按量付费":
        notes_parts.append("按量付费")
    if unit_text:
        notes_parts.append(f"单位: {unit_text}")
    if spec_info.get("spec_notes"):
        notes_parts.append(spec_info["spec_notes"])

    # ── 构建输出 ──
    engine = (engine_info.get("engine", "")
              or engine_info.get("cacheEngine", "")
              or extra_fields.get("engine", ""))

    result = {
        "sheet_name": sheet_name,
        "service": service_code,
        "instance_type": "",
        "region": region,
        "quantity": str(qty),
        "usage_hours": "720",
        "os": "Linux" if service_code == "AmazonEC2" else "",
        "engine": engine,
        "storage_gb": str(spec_info["storage_gb"]) if spec_info.get("storage_gb") else "",
        "billing_mode": item_billing_mode,
        "notes": "; ".join(notes_parts) if notes_parts else "",
        "original_request": original_request,
        "section": section,
    }

    # 应用附加字段（仅在未显式设置时）
    for k, v in extra_fields.items():
        if k not in result or not result[k]:
            result[k] = v

    # EBS 卷类型检测（从全文检测 gp3/gp2/io1 等）
    if result.get("productFamily") == "Storage" and service_code == "AmazonEBS":
        all_text_for_vol = " ".join(filter(None, [
            service_text, spec_text, description_text
        ]))
        detected_vol = detect_ebs_volume_type(all_text_for_vol)
        if detected_vol:
            result["volumeApiName"] = detected_vol
        # EBS 不需要 os 字段
        result["os"] = ""

        # 在 notes 中标注卷类型
        vol_note = f"{detected_vol}"
        if result["notes"]:
            result["notes"] = vol_note + "; " + result["notes"]
        else:
            result["notes"] = vol_note

    # 缓存引擎覆盖
    if "cacheEngine" in engine_info and service_code == "AmazonElastiCache":
        result["engine"] = engine_info["cacheEngine"]

    # 如果有 vCPU/memory 但没有 instance_type，记录到 notes 供后续推荐
    if spec_info.get("vcpu") and not result["instance_type"]:
        vcpu = spec_info["vcpu"]
        mem = spec_info.get("memory", 0)
        recommend_tag = f"recommend:{vcpu}c{mem}g"
        result["notes"] = (recommend_tag + " " + result["notes"]).strip() \
            if result["notes"] else recommend_tag
    # Memory-only managed service → also set recommend tag for Graviton resolution
    elif (spec_info.get("memory") and not result["instance_type"]
          and service_code in MANAGED_GRAVITON_SERVICES):
        mem = spec_info["memory"]
        recommend_tag = f"recommend:0c{mem}g"
        result["notes"] = (recommend_tag + " " + result["notes"]).strip() \
            if result["notes"] else recommend_tag

    # 需要实例规格的服务但未检测到任何规格 → 提醒用户
    if (service_code in SERVICES_NEED_SPEC
            and not spec_info.get("vcpu")
            and not spec_info.get("memory")
            and not spec_info.get("storage_gb")
            and not result["instance_type"]):
        no_spec_warn = "⚠️ 未指定具体规格，已使用默认最低配置查价，请确认是否符合需求"
        result["notes"] = (result["notes"] + "; " + no_spec_warn).strip("; ") \
            if result["notes"] else no_spec_warn

    # LB type annotation (ALB/NLB/CLB)
    lb_type = extra_fields.get("_lb_type")
    if lb_type and service_code == "AWSELB":
        result["notes"] = (lb_type + "; " + result["notes"]).strip("; ") \
            if result["notes"] else lb_type

    return result


def load_excel(input_path: str, region: str, billing_mode: str = "on-demand") -> list[dict]:
    """加载 Excel 文件，多 Sheet 智能解析"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要 openpyxl 库来读取 Excel: pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    all_items = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"[INFO] 处理 Sheet: '{ws_name}'", file=sys.stderr)
        items = process_sheet(ws, ws_name, region, billing_mode)
        if items:
            print(f"  → {len(items)} 条有效记录", file=sys.stderr)
            all_items.extend(items)
        else:
            print(f"  → 无有效数据", file=sys.stderr)

    wb.close()
    return all_items


# ── CSV 输入（兼容旧格式）──────────────────────────────────────────────────

def normalize_columns(row: dict) -> dict:
    """将宽松列名映射到标准列名"""
    result = {}
    for key, value in row.items():
        normalized = COLUMN_ALIASES.get(key.strip().lower(), key.strip().lower())
        result[normalized] = value
    return result


def process_csv_row(row: dict, region: str, billing_mode: str = "on-demand") -> dict | None:
    """处理 CSV 单行，返回标准化条目（兼容旧格式）"""
    norm = normalize_columns(row)

    orig_svc = norm.get("service", "")
    orig_spec = norm.get("spec", "")
    original_request = " ".join(filter(None, [orig_svc, orig_spec])).strip()

    # 处理混合计费模式（在透传之前）
    item_billing_mode = billing_mode
    if billing_mode.startswith("ri-sp-"):
        # 提取基础配置：ri-sp-1y-no-upfront -> 1y-no-upfront
        base_config = billing_mode.replace("ri-sp-", "")
        if orig_svc == "AmazonEC2":
            # EC2 使用对应的 SP
            item_billing_mode = normalize_billing_mode(f"sp-{base_config}")
        else:
            # 其他服务使用 RI
            item_billing_mode = normalize_billing_mode(f"ri-{base_config}")
    else:
        # 非混合模式：标准化简化名称
        item_billing_mode = normalize_billing_mode(billing_mode)

    # 已经是标准 ServiceCode → 透传
    known_codes = {sc for _, sc, _ in SERVICE_RULES}
    if orig_svc and orig_svc in known_codes:
        return {
            "sheet_name": "",
            "service": orig_svc,
            "instance_type": norm.get("instance_type", norm.get("spec", "")),
            "region": norm.get("region", region),
            "quantity": norm.get("quantity", "1") or "1",
            "usage_hours": norm.get("usage_hours", "720"),
            "os": norm.get("os", ""),
            "engine": norm.get("engine", ""),
            "storage_gb": norm.get("storage_gb", ""),
            "billing_mode": norm.get("billing_mode", item_billing_mode),
            "notes": norm.get("notes", ""),
            "original_request": original_request,
            "section": "",
        }

    # 需要智能匹配
    svc_text = norm.get("service", "")
    extra_text = " ".join(filter(None, [norm.get("spec", ""), norm.get("notes", "")]))
    all_text = " ".join(filter(None, [svc_text, extra_text]))
    if not all_text.strip():
        return None

    service_code, extra_fields, warning, _ = match_service_smart(svc_text)
    if not service_code:
        service_code, extra_fields, warning, _ = match_service_smart(all_text)
    if not service_code:
        service_code = svc_text.split()[0] if svc_text.strip() else "UNKNOWN"
        warning = f"无法识别服务: '{all_text.strip()}'"

    spec_text = norm.get("spec", "")
    spec_info = extract_spec(spec_text)
    for t in [svc_text, norm.get("notes", "")]:
        for k, v in extract_spec(t).items():
            if k not in spec_info:
                spec_info[k] = v

    engine_info = detect_engine(all_text)

    # 处理混合计费模式
    item_billing_mode = billing_mode
    if billing_mode.startswith("ri-sp-"):
        # 提取基础配置：ri-sp-1y-no-upfront -> 1y-no-upfront
        base_config = billing_mode.replace("ri-sp-", "")
        if service_code == "AmazonEC2":
            # EC2 使用对应的 SP
            item_billing_mode = normalize_billing_mode(f"sp-{base_config}")
        else:
            # 其他服务使用 RI
            item_billing_mode = normalize_billing_mode(f"ri-{base_config}")
    else:
        # 非混合模式：标准化简化名称
        item_billing_mode = normalize_billing_mode(billing_mode)

    result = {
        "sheet_name": "",
        "service": service_code,
        "instance_type": norm.get("instance_type", ""),
        "region": norm.get("region", region),
        "quantity": norm.get("quantity", "1") or "1",
        "usage_hours": norm.get("usage_hours", "720"),
        "os": norm.get("os", "Linux") if service_code == "AmazonEC2" else norm.get("os", ""),
        "engine": norm.get("engine", "") or engine_info.get("engine", ""),
        "storage_gb": norm.get("storage_gb", ""),
        "billing_mode": norm.get("billing_mode", item_billing_mode),
        "notes": norm.get("notes", ""),
        "original_request": original_request,
        "section": "",
    }

    for k, v in extra_fields.items():
        if not result.get(k):
            result[k] = v
    if "cacheEngine" in engine_info and service_code == "AmazonElastiCache":
        result["engine"] = engine_info["cacheEngine"]
    if spec_info.get("storage_gb") and not result["storage_gb"]:
        result["storage_gb"] = str(spec_info["storage_gb"])

    if spec_info.get("vcpu") and not result["instance_type"]:
        vcpu = spec_info["vcpu"]
        mem = spec_info.get("memory", 0)
        result["notes"] = f"recommend:{vcpu}c{mem}g" + \
            (f" {result['notes']}" if result["notes"] else "")

    if warning:
        result["notes"] = (f"{warning} " + result["notes"]).strip()

    return result


def load_csv_file(input_path: str, region: str) -> list[dict]:
    """加载 CSV 文件（兼容旧格式）"""
    items = []
    with open(input_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cleaned = {k.strip(): v.strip() for k, v in row.items()
                       if k is not None and v is not None}
            items.append(cleaned)
    return items


def load_input(input_path: str, region: str, billing_mode: str = "on-demand") -> list[dict]:
    """加载输入文件。Excel 使用智能多 Sheet 解析，CSV 使用兼容模式。"""
    path = Path(input_path)
    if path.suffix in (".xlsx", ".xls"):
        return load_excel(input_path, region, billing_mode)
    else:
        raw_rows = load_csv_file(input_path, region)
        items = []
        for i, row in enumerate(raw_rows, 1):
            result = process_csv_row(row, region, billing_mode)
            if result:
                items.append(result)
            else:
                print(f"第 {i} 行为空，已跳过", file=sys.stderr)
        return items


# ── 实例推荐（直接查询 AWS Pricing API，按价格最低推荐）──────────────────

_ec2_instance_cache: dict[str, list[dict]] = {}


def _query_all_ec2_instances(region: str) -> list[dict]:
    """查询区域内所有 EC2 Linux/Shared 实例价格，结果缓存避免重复调用"""
    if region in _ec2_instance_cache:
        return _ec2_instance_cache[region]

    sys.path.insert(0, str(Path(__file__).parent))
    from query_price import run_aws_cli, extract_pricing

    filters = [
        {"Type": "TERM_MATCH", "Field": "regionCode", "Value": region},
        {"Type": "TERM_MATCH", "Field": "operatingSystem", "Value": "Linux"},
        {"Type": "TERM_MATCH", "Field": "tenancy", "Value": "Shared"},
        {"Type": "TERM_MATCH", "Field": "capacitystatus", "Value": "Used"},
        {"Type": "TERM_MATCH", "Field": "preInstalledSw", "Value": "NA"},
    ]

    instances = []
    seen_types = set()
    next_token = None

    print(f"[INFO] 正在查询 EC2 实例价格 ({region})...", file=sys.stderr)

    for _ in range(10):  # 最多 10 页
        args = [
            "pricing", "get-products",
            "--service-code", "AmazonEC2",
            "--filters", json.dumps(filters),
            "--max-results", "100",
        ]
        if next_token:
            args += ["--next-token", next_token]

        data = run_aws_cli(args, timeout=60)
        if not data:
            break

        for price_str in data.get("PriceList", []):
            product = json.loads(price_str) if isinstance(price_str, str) else price_str
            attrs = product.get("product", {}).get("attributes", {})

            instance_type = attrs.get("instanceType", "")
            if not instance_type or instance_type in seen_types:
                continue
            seen_types.add(instance_type)

            try:
                vcpu = int(attrs.get("vcpu", "0"))
                mem_str = (attrs.get("memory", "0")
                           .replace(",", "").replace(" GiB", "")
                           .replace(" GB", "").strip())
                memory = float(mem_str)
            except (ValueError, TypeError):
                continue

            pricing = extract_pricing(product)
            od = pricing.get("on_demand")
            if not od or od["price"] == "N/A":
                continue

            try:
                hourly_price = float(od["price"])
            except (ValueError, TypeError):
                continue
            if hourly_price <= 0:
                continue

            instances.append({
                "instance_type": instance_type,
                "vcpu": vcpu,
                "memory": memory,
                "hourly_price": hourly_price,
            })

        next_token = data.get("NextToken")
        if not next_token:
            break

    print(f"[INFO] 获取到 {len(instances)} 种 EC2 实例类型", file=sys.stderr)
    _ec2_instance_cache[region] = instances
    return instances


def _is_graviton(instance_type: str) -> bool:
    """判断实例类型是否为 Graviton (ARM) 架构"""
    family = instance_type.split(".")[0] if "." in instance_type else instance_type
    return bool(re.match(r'^[a-z]+\d+g', family))


def _find_cheapest_instance(instances: list[dict],
                            vcpu_need: int, memory_need: float,
                            exclude_families: list[str] | None = None,
                            arch: str = "x86") -> dict | None:
    """找到满足 vCPU >= 需求 AND memory >= 需求的最优实例。

    根据 CPU:内存比例智能选择实例族：
    - 1:1~1:2 → 计算优化 (c 系列)
    - 1:2~1:4 → 通用 (m 系列)
    - 1:4+    → 内存优化 (r 系列)

    在匹配的系列内按价格最低选择。如果目标系列没有匹配，回落到所有系列中价格最低的。
    """
    candidates = [i for i in instances
                  if i["vcpu"] >= vcpu_need and i["memory"] >= memory_need]

    if exclude_families:
        candidates = [i for i in candidates
                      if not any(i["instance_type"].startswith(f + ".")
                                 for f in exclude_families)]

    if arch == "x86":
        candidates = [i for i in candidates
                      if not _is_graviton(i["instance_type"])]
    elif arch == "arm":
        candidates = [i for i in candidates
                      if _is_graviton(i["instance_type"])]

    if not candidates:
        return None

    # 根据 CPU:内存比例确定目标实例族
    if vcpu_need > 0 and memory_need > 0:
        ratio = memory_need / vcpu_need
        if ratio < 2:
            preferred_prefix = "c"  # 计算优化 (1:1~1:2, 严格小于2)
        elif ratio <= 4:
            preferred_prefix = "m"  # 通用 (1:2~1:4, ratio=2 归 m 系列)
        else:
            preferred_prefix = "r"  # 内存优化
    else:
        preferred_prefix = "m"  # 默认通用

    # 在目标系列内找最便宜的
    preferred = [i for i in candidates
                 if i["instance_type"].split(".")[0].startswith(preferred_prefix)]

    if preferred:
        return min(preferred, key=lambda i: i["hourly_price"])

    # 目标系列没有匹配，回落到所有候选中最便宜的
    return min(candidates, key=lambda i: i["hourly_price"])


def _resolve_managed_graviton_instance(service: str, vcpu: int = 0,
                                       memory: int = 0) -> tuple[str, str]:
    """为托管服务选择最新代 Graviton 实例规格。

    返回 (instance_type, description)。无匹配时返回 ("", "")。
    中国区各服务可用的 Graviton 实例族不同（如 DocumentDB 仅 r6g），
    通过 MANAGED_GRAVITON_FAMILY 查表确定。
    """
    prefix = MANAGED_SERVICE_PREFIX.get(service)
    if prefix is None:
        return "", ""

    suffix = ".search" if service == "AmazonES" else ""

    # 查服务级别的 Graviton 实例族覆盖
    svc_families = MANAGED_GRAVITON_FAMILY.get(service)

    def _get_family(series: str) -> tuple[str, list]:
        """获取指定系列(r/m)的实际可用 Graviton 族和规格表"""
        if svc_families and series in svc_families:
            return svc_families[series]
        # 默认: r7g / m7g
        if series == "r":
            return "r7g", GRAVITON_R7G_SPECS
        return "m7g", GRAVITON_M7G_SPECS

    def _build_type(family: str, size: str) -> str:
        if prefix:
            return f"{prefix}.{family}.{size}"
        return f"{family}.{size}{suffix}"

    if vcpu > 0 and memory > 0:
        # Both specified: r series if memory >= vcpu*4, else m series
        if memory >= vcpu * 4:
            family, specs = _get_family("r")
        else:
            family, specs = _get_family("m")
        for sv, sm, sname in specs:
            if sv >= vcpu and sm >= memory:
                inst = _build_type(family, sname)
                return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB)"
        # Fallback: largest available
        sv, sm, sname = specs[-1]
        inst = _build_type(family, sname)
        return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB, 最大可用)"
    elif memory > 0:
        # Memory-only: prefer r series (memory-optimized)
        family, specs = _get_family("r")
        for sv, sm, sname in specs:
            if sm >= memory:
                inst = _build_type(family, sname)
                return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB)"
        sv, sm, sname = specs[-1]
        inst = _build_type(family, sname)
        return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB, 最大可用)"
    elif vcpu > 0:
        # vCPU-only: use m series (general purpose)
        family, specs = _get_family("m")
        for sv, sm, sname in specs:
            if sv >= vcpu:
                inst = _build_type(family, sname)
                return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB)"
        sv, sm, sname = specs[-1]
        inst = _build_type(family, sname)
        return inst, f"Graviton {family}.{sname} ({sv}vCPU/{sm}GiB, 最大可用)"

    return "", ""


def resolve_instance_recommendations(items: list[dict], region: str) -> list[dict]:
    """对含有 recommend:XcYg 的条目，自动推荐匹配实例。

    EC2: 查询 Pricing API，推荐最新代 Intel（排除 t 系列和 Graviton）。
    托管服务 (RDS/ElastiCache/等): 直接映射到最新代 Graviton 实例规格。
    """
    # 仅在有 EC2 推荐需求时查询 API（一次查询覆盖所有规格）
    need_ec2 = any(
        re.search(r'recommend:(\d+)c(\d+)g', item.get("notes", ""))
        and item["service"] == "AmazonEC2"
        for item in items
    )
    instances = _query_all_ec2_instances(region) if need_ec2 else []
    rec_cache: dict[tuple, dict | None] = {}

    for item in items:
        notes = item.get("notes", "")
        m = re.search(r'recommend:(\d+)c(\d+)g', notes)
        if not m:
            continue

        vcpu = int(m.group(1))
        memory = int(m.group(2))
        service = item["service"]

        if service != "AmazonEC2":
            if service in MANAGED_GRAVITON_SERVICES:
                inst, desc = _resolve_managed_graviton_instance(
                    service, vcpu, memory
                )
                if inst:
                    item["instance_type"] = inst
                    # 超配备注：仅在实际规格大于需求时标注
                    mem_match = re.search(r'/(\d+)GiB', desc)
                    if mem_match and memory > 0:
                        actual_mem = int(mem_match.group(1))
                        if actual_mem > memory:
                            rec_desc = f"最小匹配规格 {inst}({actual_mem}GiB)"
                        else:
                            rec_desc = ""
                    else:
                        rec_desc = ""
                    item["notes"] = notes.replace(
                        m.group(0), rec_desc
                    ).strip()
                else:
                    item["notes"] = notes.replace(
                        m.group(0), f"需要{vcpu}vCPU/{memory}GiB"
                    ).strip()
            else:
                item["notes"] = notes.replace(
                    m.group(0), f"需要{vcpu}vCPU/{memory}GiB"
                ).strip()
            continue

        # 检测用户是否要求 t 系列或 ARM/Graviton
        context = (notes + " " + item.get("original_request", "")).lower()

        want_t = any(kw in context for kw in ["t2", "t3", "t4g", "突发", "burstable"])
        want_amd = any(kw in context for kw in ["amd", "c5a", "m5a", "r5a", "c6a", "m6a", "r6a"])
        want_old = any(kw in context for kw in ["c5", "m5", "r5", "c4", "m4", "r4"])
        exclude_families = []
        if not want_t:
            exclude_families += ["t2", "t3", "t3a", "t4g"]
        if not want_amd:
            exclude_families += ["c5a", "c5ad", "m5a", "m5ad", "r5a", "r5ad",
                                "c6a", "m6a", "r6a", "c7a", "m7a", "r7a",
                                "hpc7a"]
        if not want_old:
            # 排除旧代 Intel（默认用最新代）
            exclude_families += ["c4", "c5", "c5d", "c5n",
                                "m4", "m5", "m5d", "m5n", "m5dn",
                                "r4", "r5", "r5d", "r5n", "r5dn",
                                "i3", "i3en", "d2", "h1",
                                "p3", "g3", "g3s", "x1", "x1e", "z1d"]

        want_arm = any(kw in context for kw in ["arm", "graviton"])
        arch = "arm" if want_arm else "x86"

        # 按 (vcpu, memory, exclude_families, arch) 缓存推荐结果
        key = (vcpu, memory, tuple(exclude_families), arch)
        if key not in rec_cache:
            rec_cache[key] = _find_cheapest_instance(
                instances, vcpu, memory,
                exclude_families=exclude_families, arch=arch
            )

        best = rec_cache[key]
        if best:
            item["instance_type"] = best["instance_type"]
            if memory > 0 and best["memory"] > memory:
                rec_desc = (f"最小匹配规格 {best['instance_type']}"
                            f"({best['memory']:.0f}GiB)")
            else:
                rec_desc = ""
            item["notes"] = notes.replace(m.group(0), rec_desc).strip()
        else:
            item["notes"] = notes.replace(
                m.group(0), f"需手动选择实例({vcpu}vCPU/{memory}GiB)"
            ).strip()

    return items


# ── 输出 ──────────────────────────────────────────────────────────────────

def save_csv(items: list[dict], output_path: str):
    """保存标准化 CSV"""
    if not items:
        print("没有数据可输出", file=sys.stderr)
        return

    fieldnames = list(OUTPUT_FIELDS)
    for item in items:
        for k in item:
            if k not in fieldnames:
                fieldnames.append(k)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(items)

    print(f"[OK] 已输出标准化 CSV: {output_path} ({len(items)} 条)", file=sys.stderr)


def print_csv(items: list[dict]):
    """输出标准化 CSV 到 stdout"""
    if not items:
        return
    writer = csv.DictWriter(sys.stdout, fieldnames=OUTPUT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(items)


def _output_excel(items: list[dict], args):
    """计算成本并生成 Excel 报价单（一步到位）"""
    sys.path.insert(0, str(Path(__file__).parent))
    from calculate_cost import (
        load_discount_config, get_price_for_item, calculate_item_cost,
    )
    from generate_quote import generate_quote

    discount_config = load_discount_config(args.discount_config)

    # 设置默认区域
    for item in items:
        if not item.get("region"):
            item["region"] = args.region

    # 逐条查询价格并计算成本
    print("[INFO] 正在查询价格并计算成本...", file=sys.stderr)
    results = []
    for item in items:
        billing_mode = item.get("billing_mode", "on-demand") or "on-demand"
        print(f"  查询 {item.get('service', '')} {item.get('instance_type', '')} ...",
              file=sys.stderr)
        price_data = get_price_for_item(item, billing_mode=billing_mode)
        if not price_data:
            results.append({
                "service": item.get("service", ""),
                "instance_type": item.get("instance_type", ""),
                "region": item.get("region", ""),
                "quantity": int(item.get("quantity", 1) or 1),
                "usage_hours": float(item.get("usage_hours", 720) or 720),
                "billing_mode": billing_mode,
                "hourly_list": 0, "hourly_after_discount": 0,
                "monthly_per_unit": 0, "monthly_total": 0,
                "upfront_total": 0, "yearly_total": 0,
                "applied_discounts": [],
                "notes": item.get("notes", ""),
                "original_request": item.get("original_request", ""),
                "currency": "CNY",
                "sheet_name": item.get("sheet_name", ""),
                "section": item.get("section", ""),
            })
            continue
        cost = calculate_item_cost(item, price_data, discount_config, args.include_tax)
        results.append(cost)

    # 生成报价单
    gen_config = {
        "customer": args.customer,
        "validity": 30,
        "include_tax": args.include_tax,
    }
    wb = generate_quote(items, results, gen_config)
    wb.save(args.output)

    # 打印摘要
    total_monthly = sum(r.get("monthly_total", 0) for r in results)
    total_yearly = sum(r.get("yearly_total", 0) for r in results)
    tax_label = "（含税）" if args.include_tax else "（不含税）"
    print(f"\n[OK] 报价单已生成: {args.output}", file=sys.stderr)
    print(f"\n摘要{tax_label}:", file=sys.stderr)
    print(f"  条目数: {len(results)}", file=sys.stderr)
    print(f"  月费用合计: CNY {total_monthly:,.2f}", file=sys.stderr)
    print(f"  年费用合计: CNY {total_yearly:,.2f}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="AWS 中国区智能工作负载导入 v2 - 支持任意格式 Excel 多 Sheet"
    )
    parser.add_argument("--input", "-i", required=True,
                        help="输入 CSV/Excel 文件路径")
    parser.add_argument("--output", "-o",
                        help="输出标准化 CSV 路径（不指定则输出到 stdout）")
    parser.add_argument("--region", "-r", default="cn-north-1",
                        help="默认区域 (默认: cn-north-1)")
    parser.add_argument("--calculate", action="store_true",
                        help="预处理后直接调用 calculate_cost.py 计算")
    parser.add_argument("--profile", default=None,
                        help="AWS CLI profile (默认: 不指定则用 AWS CLI 默认配置)")
    parser.add_argument("--no-recommend", action="store_true",
                        help="跳过实例推荐（加速处理）")
    parser.add_argument("--discount-config", "-d",
                        default=str(Path(__file__).parent.parent / "discount-config.yaml"),
                        help="折扣配置文件路径")
    parser.add_argument("--include-tax", action="store_true",
                        help="含 6%% 增值税")
    parser.add_argument("--customer", default="",
                        help="客户名称（报价单用）")
    parser.add_argument("--billing-mode", "-b", default="on-demand",
                        choices=["on-demand", "ri-1y-no-upfront", "ri-1y-partial", "ri-1y-all",
                                "ri-3y-no-upfront", "ri-3y-partial", "ri-3y-all",
                                "sp-1y-no-upfront", "sp-1y-partial", "sp-1y-all",
                                "sp-3y-no-upfront", "sp-3y-partial", "sp-3y-all",
                                "ri-sp-1y-no-upfront", "ri-sp-1y-partial", "ri-sp-1y-all",
                                "ri-sp-3y-no-upfront", "ri-sp-3y-partial", "ri-sp-3y-all"],
                        help="计费模式 (默认: on-demand)")
    args = parser.parse_args()

    # 默认输出 Excel 报价单
    if not args.output and not args.calculate:
        input_path = Path(args.input)
        args.output = str(input_path.parent / f"{input_path.stem}_报价单.xlsx")

    # 设置 AWS profile
    if args.profile:
        sys.path.insert(0, str(Path(__file__).parent))
        import query_price
        query_price.AWS_PROFILE = args.profile

    items = load_input(args.input, args.region, args.billing_mode)
    if not items:
        print("[ERROR] 输入文件为空或无有效数据", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] 共解析 {len(items)} 条有效记录", file=sys.stderr)

    if not args.no_recommend:
        items = resolve_instance_recommendations(items, args.region)

    # 输出映射摘要
    print("\n[映射结果]", file=sys.stderr)
    current_sheet = None
    current_section = None
    for i, item in enumerate(items, 1):
        s = item.get("sheet_name", "")
        sec = item.get("section", "")
        if s and s != current_sheet:
            print(f"\n  ── Sheet: {s} ──", file=sys.stderr)
            current_sheet = s
            current_section = None
        if sec and sec != current_section:
            print(f"  [{sec}]", file=sys.stderr)
            current_section = sec

        svc = item["service"]
        inst = item.get("instance_type", "")
        qty = item.get("quantity", "1")
        extra = []
        if item.get("engine"):
            extra.append(f"engine={item['engine']}")
        if item.get("storage_gb"):
            extra.append(f"storage={item['storage_gb']}GB")
        if inst:
            extra.append(f"instance={inst}")
        extra_str = f" ({', '.join(extra)})" if extra else ""
        print(f"  {i}. {svc} x{qty}{extra_str}", file=sys.stderr)
    print("", file=sys.stderr)

    if args.calculate:
        import tempfile
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8"
        ) as tmp:
            tmp_path = tmp.name
            writer = csv.DictWriter(
                tmp, fieldnames=OUTPUT_FIELDS, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(items)

        script_dir = Path(__file__).parent
        calc_script = script_dir / "calculate_cost.py"
        cmd = [sys.executable, str(calc_script),
               "--input", tmp_path, "--region", args.region]
        if args.profile:
            cmd += ["--profile", args.profile]
        print(f"[INFO] 调用: {' '.join(cmd)}", file=sys.stderr)
        result = subprocess.run(cmd)
        Path(tmp_path).unlink(missing_ok=True)
        sys.exit(result.returncode)

    elif args.output and args.output.endswith(".xlsx"):
        _output_excel(items, args)
    elif args.output:
        save_csv(items, args.output)
    else:
        print_csv(items)


if __name__ == "__main__":
    main()
